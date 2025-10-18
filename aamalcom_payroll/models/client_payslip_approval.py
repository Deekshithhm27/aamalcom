# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from datetime import datetime, date
from odoo.exceptions import UserError, ValidationError
import io
import base64
import calendar

# Import openpyxl for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image
except ImportError:
    openpyxl = None
    # No logger defined, keep the exception handling clean


class ClientPayslipApproval(models.Model):
    _name = 'client.payslip.approval'
    _description = "Client Payroll Approval"
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char(
        string="Payslip Reference",
        required=True,
        copy=False,
        readonly=True,
        default=lambda self: self.env['ir.sequence'].next_by_code('client.payslip.approval')
    )
    active = fields.Boolean('Active', default=True)
    user_id = fields.Many2one('res.users', string='User', default=lambda self: self.env.user)
    company_id = fields.Many2one('res.company', string='Company', required=True, default=lambda self: self.env.user.company_id)
    client_parent_id = fields.Many2one('res.partner',string="Client",domain=[('is_company', '=', True), ('parent_id', '=', False)])
    client_id = fields.Many2one('res.users',string="Client SPOC",related="client_parent_id.client_id", store=True)

    from_date = fields.Date(string="From Date", required=True, tracking=True)
    to_date = fields.Date(string="To Date", required=True, tracking=True)
    create_date = fields.Datetime(string="Create Date", readonly=True, default=lambda self: datetime.now())
    processed_date = fields.Datetime(string="Processed Date", readonly=True, tracking=True)
    
    approval_payslip_document = fields.Binary(string="External Payroll Document")
    approval_payslip_filename = fields.Char(string="Document Filename")
    
    generated_payroll_document = fields.Binary(string="Generated Payroll Document")
    generated_payroll_filename = fields.Char(string="Generated Filename")
    
    payslip_documents_ids = fields.One2many(
        comodel_name='client.payslip.document',
        inverse_name='payslip_approval_id',
        string='Updated Documents'
    )
    
    refusal_reason = fields.Text(string="Refusal Reason", tracking=True, readonly=True)

    state = fields.Selection([
        ('draft', 'Draft'),
        ('submit_to_payroll', 'Submitted for Verification'),
        ('submit_to_payroll_employee', 'Submitted to Payroll Employee'),
        ('submit_to_pm', 'Submitted to PM'),
        ('approved_by_pm', 'Verified By PM'),
        ('submit_to_fm', 'Submitted to Finance Manager'),
        ('submit_to_payroll_manager', 'Submitted to Payroll Manager'),
        ('done', 'Completed'),
        ('refuse', 'Refused'),
    ], string="Status", default='draft', tracking=True)

    is_payroll_manager = fields.Boolean(
        string="Is Payroll Manager?",
        compute="_compute_is_payroll_manager"
    )
    disbursed_date = fields.Date(
        string='Payroll Disbursment Date',
        readonly=True, 
        states={'submit_to_payroll_employee': [('readonly', False)]}
    )
    
    approval_document_date = fields.Date(string="Approval Document Date", compute="_compute_approval_document_date", store=True)
    creator_pm_id = fields.Many2one(
            'res.users',
            string="Creating Project Manager",
            readonly=True,
            tracking=True,
            help="Stores the user who created this record, only if they are a Project Manager."
        )
    notification_sent = fields.Boolean(
            string="Notification Sent",
            default=False,
            readonly=True,
            copy=False,
            tracking=True
        )
    
    @api.model
    def create(self, vals):
        record = super(ClientPayslipApproval, self).create(vals) 
        # Check if the creator belongs to the PM group
        pm_group_ext_id = 'visa_process.group_service_request_manager'

        if self.env.user.has_group(pm_group_ext_id):
            record.creator_pm_id = self.env.user.id
                
        return record

    @api.depends('approval_payslip_document')
    def _compute_approval_document_date(self):
        for rec in self:
            if rec.approval_payslip_document and not rec.approval_document_date:
                rec.approval_document_date = date.today()
            elif not rec.approval_payslip_document:
                rec.approval_document_date = False
    
    def _get_related_salary_records(self):
        """Helper function to find the relevant salary tracking records."""
        return self.env['client.emp.salary.tracking'].search([
            ('client_parent_id', '=', self.client_parent_id.id),
            ('date_start', '>=', self.from_date),
            ('date_end', '<=', self.to_date),
            ])
    
    @api.depends()
    def _compute_is_payroll_manager(self):
        """Check if the current user belongs to the payroll manager group."""
        for rec in self:
            rec.is_payroll_manager = self.env.user.has_group('visa_process.group_service_request_payroll_employee')
    
    def _get_bank_details_from_service_request(self, employee_tracking_record):
        """
        Fetches Bank and IBAN using the confirmed field names (bank_id and bic)
        based on the employee's service request type.
        """
        bank_name = ''
        iban_number = ''
        
        service_type = employee_tracking_record.service_request_type
        employee_id = employee_tracking_record.employee_id.id
        
        if service_type == 'employment_visa':
            # Search for the relevant Employment Visa record
            visa_record = self.env['employment.visa'].search([
                ('employee_name', '=', employee_id),
                ('state', '=', 'approved'),
            ], limit=1, order='create_date desc')
            
            if visa_record:
                bank_name = visa_record.bank_id.name or ''
                iban_number = visa_record.bic or ''
                
        elif service_type == 'local_transfer':
            # Search for the relevant Local Transfer record
            transfer_record = self.env['local.transfer'].search([
                ('employee_name', '=', employee_id),
                ('state', '=', 'approved'),
            ], limit=1, order='create_date desc')
            
            if transfer_record:
                bank_name = transfer_record.bank_id.name or ''
                iban_number = transfer_record.bic or ''
                
        return bank_name, iban_number
    
    def action_generate_payroll_xsl(self):
            self.ensure_one()

            if not openpyxl:
                raise UserError(_("The 'openpyxl' Python library is required to generate the Excel file. Please install it."))

            # 1. Prepare Data and Formatting
            company = self.env.company
            client_name = self.client_parent_id.name or "Client"
            
            month_name = calendar.month_abbr[self.from_date.month].upper()
            year = self.from_date.year
            month_year_str = f"{month_name} - {year}"

            # 2. Search for salary tracking records
            domain = [
                ('client_parent_id', '=', self.client_parent_id.id),
                ('date_start', '>=', self.from_date),
                ('date_end', '<=', self.to_date),
                ('state', '=', 'draft'),
            ]

            salary_records = self.env['client.emp.salary.tracking'].search(domain)

            if not salary_records:
                raise UserError(_("No employee salary tracking records in 'Draft' found for the selected client and period."))

            # 3. Generate XSL (Excel) file
            output = io.BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Payroll"

            # Define Styles
            header_font = Font(name='Arial', size=16, bold=True)
            company_font = Font(name='Arial', size=10, bold=True)
            table_header_font = Font(name='Arial', size=10, bold=True)
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            # Set Column Widths (A-R)
            column_widths = {'A': 5, 'B': 15, 'C': 25, 'D': 15, 'E': 30, 'F': 10, 
                             'G': 10, 'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 10, 
                             'M': 10, 'N': 10, 'O': 10, 'P': 10, 'Q': 10, 'R': 10}
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            # A. Insert Company Logo
            if company.logo:
                try:
                    logo_data = base64.b64decode(company.logo)
                    logo_image = Image(io.BytesIO(logo_data))
                    
                    logo_image.width = 150
                    logo_image.height = 100
                    
                    sheet.add_image(logo_image, 'A1')
                    sheet.row_dimensions[1].height = 60
                except Exception as e:
                    # Removed logger
                    pass
            
            # B. Write Company Name Header
            sheet.merge_cells('C1:R1')
            sheet['C1'].value = company.name or 'Aamal Com Company'
            sheet['C1'].font = company_font
            sheet['C1'].alignment = Alignment(horizontal='right', vertical='center')
            
            # C. Write Payroll Title
            title_row = 16
            sheet.merge_cells(start_row=title_row, start_column=3, end_row=title_row, end_column=18)
            sheet[f'C{title_row}'].value = "Aamal Com Company"
            sheet[f'C{title_row}'].font = header_font
            sheet[f'C{title_row}'].alignment = center_align
            
            # D. Write Sub-Header
            sub_title_row = 17
            sheet.merge_cells(start_row=sub_title_row, start_column=3, end_row=sub_title_row, end_column=18)
            sub_title_value = f"Internal Payroll - {client_name} - MONTH - {month_year_str}"
            sheet[f'C{sub_title_row}'].value = sub_title_value
            sheet[f'C{sub_title_row}'].font = Font(name='Arial', size=12, bold=True)
            sheet[f'C{sub_title_row}'].alignment = center_align

            # E. Table Headers
            table_start_row = 19
            
            employee_headers = [
                "SL No", "Employee ID", "Employee Name", "Iqama Number", "IBAN", "BANK NAME", "Days In Month", "Days Present"
            ]
            
            salary_headers = [
                "Basic Salary", "Housing Allowance", "Transpo Allowance", "Gross Salary", 
                "Other Allowance", "Over Time Allowance", "Additions", "Gross Salary"
            ]

            # Write Employee Headers
            for col_idx, header in enumerate(employee_headers, start=1):
                cell = sheet.cell(row=table_start_row, column=col_idx, value=header)
                cell.font = table_header_font
                cell.alignment = center_align
                cell.border = thin_border
                sheet.row_dimensions[table_start_row].height = 30
                
            # Write Salary Headers
            for col_idx, header in enumerate(salary_headers, start=len(employee_headers) + 1):
                cell = sheet.cell(row=table_start_row, column=col_idx, value=header)
                cell.font = table_header_font
                cell.alignment = center_align
                cell.border = thin_border

            # F. Populate Data Rows
            data_row_start = table_start_row + 1
            
            for index, record in enumerate(salary_records, start=1):
                bank_name, iban_number = self._get_bank_details_from_service_request(record)
                row = [
                    index,
                    record.client_emp_sequence or '',
                    record.employee_id.name or '',
                    record.employee_id.iqama_no or '',
                    iban_number,
                    bank_name,
                    self.to_date.day,
                    record.worked_days,
                    
                    record.wage,
                    record.hra,
                    record.travel_allowance,
                    record.gross_salary,
                    record.other_allowance,
                    record.overtime,
                    record.additions,
                    record.gross_salary,
                ]
                
                # Append data row and apply border
                sheet.append(row)
                for col_idx in range(1, len(row) + 1):
                    cell = sheet.cell(row=data_row_start + index - 1, column=col_idx)
                    cell.border = thin_border
                    # Format money columns (J, K, L, M, O, P, Q, R are columns 9 onwards)
                    if col_idx >= 9:
                        cell.number_format = '#,##0.00'


            # 4. Save and Attach the file
            workbook.save(output)
            output.seek(0)
            excel_data = output.read()

            client_name_safe = client_name.replace(' ', '_').replace('/', '_')
            date_str = self.from_date.strftime('%Y%m')
            filename = "Generated_Payroll_%s_%s.xlsx" % (client_name_safe, date_str)

            self.write({
                'generated_payroll_document': base64.b64encode(excel_data),
                'generated_payroll_filename': filename,
            })
            
            # 5. Update the state of the salary tracking records
            salary_records.write({'state': 'submit_to_payroll'})
            self.message_post(body=_("Payroll file '%s' successfully generated and attached. %d employee records moved to 'Submit to Payroll'.") % (filename, len(salary_records)))

            # 6. Return an action to force the client to reload the form
            return {
                'type': 'ir.actions.client',
                'tag': 'reload',
            }
    @api.constrains('from_date', 'to_date')
    def _check_current_month_period(self):
        """Constraint to ensure the period is close to the current date,
        e.g., within the current month/year, to prevent future payroll runs.
        """
        today = date.today()
        current_month_start = today.replace(day=1)
            
        for run in self:
            #  Restrict strictly to the current month
            if run.from_date.year > today.year or (run.from_date.year == today.year and run.from_date.month > today.month):
                raise ValidationError(_("You cannot create a payroll batch for a future month."))
                     
            if run.to_date.year > today.year or (run.to_date.year == today.year and run.to_date.month > today.month):
                raise ValidationError(_("The payroll period cannot end in a future month."))
                    
            # to ensure the date_start is before date_end
            if run.from_date and run.to_date and run.from_date > run.to_date:
                raise ValidationError(_("The start date must be before the end date."))
    
    def action_submit_to_payroll(self):
        for rec in self:
            # 1. Get the current user's employee record
            current_user_employee = self.env['hr.employee'].search([
                ('user_id', '=', self.env.user.id)
            ], limit=1)
            
            # Validation 1: Check if the current user is linked to an employee.
            if not current_user_employee:
                raise UserError(_(
                    "Action Denied: You must be linked to an Employee record to submit this request."
                ))

            # 2. Get the assigned Project Manager from the Client's SPOC
            client_pm_employee = rec.client_parent_id.company_spoc_id
            
            # Validation 2: Check if the current user matches the client's assigned PM.
            if client_pm_employee and client_pm_employee.id != current_user_employee.id:
                raise UserError(_(
                    "Authorization Error: You are not the Project Manager ('%s') assigned to Client '%s'. "
                    "Only the assigned Project Manager can submit this payroll."
                ) % (client_pm_employee.name, rec.client_parent_id.name))
            
            if not rec.client_parent_id:
                raise UserError(_("Please select a Client before submitting the payroll."))
                
            # If validation passes:
            rec.state = 'submit_to_payroll'
            rec.processed_date = datetime.now()
                
        return True

    def action_submit_to_pm(self): 
        if not self.generated_payroll_document:
            raise UserError(_("Please generate payroll first before submitting for verification."))
        if not self.payslip_documents_ids:
            raise UserError(_("Please upload at least one document in the 'Documents' session before submitting for approval."))
        
        self.write({'state': 'submit_to_pm', 'processed_date': datetime.now()})
        self._get_related_salary_records().write({'state': 'submit_to_pm'})

        return True

    def action_verified_by_pm(self):          
        self.write({'state': 'approved_by_pm', 'processed_date': datetime.now()})
        self._get_related_salary_records().write({'state': 'approved_by_pm'})

        return True
    
    def action_submit_to_fm_manager(self):
        if not self.payslip_documents_ids:
            raise UserError(_("Please upload at least one document in the 'Documents' session before submitting for approval."))
                     
        self.write({'state': 'submit_to_fm', 'processed_date': datetime.now()})
        self._get_related_salary_records().write({'state': 'submit_to_fm'})

        return True

    def action_submit_to_payroll_manager(self):
        if not self.payslip_documents_ids:
            raise UserError(_("Please upload at least one document in the 'Documents' session before submitting for approval."))
                     
        self.write({'state': 'submit_to_payroll_manager', 'processed_date': datetime.now()})
        self._get_related_salary_records().write({'state': 'submit_to_payroll_manager'})

        return True

    def action_submit_to_payroll_employee(self):
        self.write({'state': 'submit_to_payroll_employee', 'processed_date': datetime.now()})
        self._get_related_salary_records().write({'state': 'submit_to_payroll_employee'})

        return True

    def action_reviewed_by_payroll_employee(self):
        self.write({'state': 'done', 'processed_date': datetime.now()})
        self._get_related_salary_records().write({'state': 'done'})

        return True

    def action_refused_by_pm(self):
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'client.payslip.refuse.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_payslip_approval_id': self.id,
                }
            }
    
    def action_re_submit(self):
        for rec in self:
            rec.state = 'submit_to_payroll'
            rec.processed_date = datetime.now()
        return True
    
    def action_send_disbursement_notification(self):
        self.ensure_one()

        # 1. Get the assigned Project Manager (SPOC) from the Client record
        recipient_employee = self.client_parent_id.company_spoc_id
        
        recipient_user = False
        recipient_name = "Unknown Recipient"
        
        if recipient_employee and recipient_employee.user_id:
            # Case 1: Client SPOC (Employee) is set and linked to a User record (res.users)
            recipient_user = recipient_employee.user_id
            recipient_name = recipient_employee.name
        else:
            # Case 2: Fallback to the record's creator or PM
            recipient_user = self.creator_pm_id or self.create_uid
            if recipient_user:
                recipient_name = recipient_user.name

        # 2. Final check for a valid recipient user
        if not recipient_user:
            raise UserError(_("Cannot send notification: The Client SPOC is missing or not linked to a user, and no record creator was found."))

        # 3. Get the Partner record for email validation
        recipient_partner = recipient_user.partner_id

        # 4. Email validation
        if not recipient_partner or (not recipient_partner.email and not recipient_partner.work_email):
            raise UserError(_("The intended recipient (%s) does not have a work email or standard email configured.") % (recipient_name,))

        # 5. Send the email
        template = self.env.ref('aamalcom_payroll.email_template_payroll_disbursement_notification')
        
        template.send_mail(self.id, force_send=True)
        
        # 6. Log and update state
        self.write({'notification_sent': True})
        self.message_post(body=_("Disbursement notification email successfully triggered to Client SPOC: %s.") % (recipient_name))
        
        return True


class ClientPayslipRefuseWizard(models.TransientModel):
    _name = 'client.payslip.refuse.wizard'
    _description = 'Wizard to Refuse Payslip Approval'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    reason = fields.Text(string="Reason for Refusal", required=True)    
    payslip_approval_id = fields.Many2one('client.payslip.approval', string="Payslip Approval")

    def refuse_payslip(self):
        self.ensure_one()
        payslip = self.payslip_approval_id
            
        payslip.state = 'refuse'
        payslip.refusal_reason = self.reason
            
        payslip.message_post(body="Payslip has been refused by PM with the following reason: <br/>%s" % self.reason)

        return {'type': 'ir.actions.act_window_close'}